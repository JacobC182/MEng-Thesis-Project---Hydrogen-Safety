#ML tuning results analysis script

#importing libraries
import numpy as np
import matplotlib.pyplot as plt
from numpy.lib import index_exp
import openpyxl

#Importing results function
def Import():

    RF_results = []
    XT_results = []
    GB_results = []
    KN_results = []
    SV_results = []
    SVP_results = []
    MP_results = []

    WB = openpyxl.load_workbook(filename="CV-Results.xlsx", read_only=True)
    S = WB.active

    for row in S.iter_rows(min_row=3, max_row=121, min_col=2, max_col=4, values_only=True):
        RF_results.append(list(row))

    for row in S.iter_rows(min_row=3, max_row=121, min_col=6, max_col=8, values_only=True):
        XT_results.append(list(row))
    
    for row in S.iter_rows(min_row=3, max_row=2162, min_col=10, max_col=14, values_only=True):
        GB_results.append(list(row))
    
    for row in S.iter_rows(min_row=3, max_row=1002, min_col=16, max_col=20, values_only=True):
        KN_results.append(list(row))
    
    for row in S.iter_rows(min_row=3, max_row=272, min_col=22, max_col=26, values_only=True):
        SV_results.append(list(row))

    for row in S.iter_rows(min_row=3, max_row=362, min_col=28, max_col=32, values_only=True):
        SVP_results.append(list(row))
    
    for row in S.iter_rows(min_row=3, max_row=7619, min_col=34, max_col=37, values_only=True):
        MP_results.append(list(row))

    return RF_results, XT_results, GB_results, KN_results, SV_results, SVP_results, MP_results

#Extracting results data
RF_results, XT_results, GB_results, KN_results, SV_results, SVP_results, MP_results = Import()

#Random Forest
def RF_analysis():

    RF_array = np.array(RF_results)
    Best_RF = np.max(RF_array[:,1])
    Low_RF = np.min(RF_array[:,2])

    RF_Best_Index = np.where(RF_array[:,1]==Best_RF)[0][0]
    RF_Low_Index = np.where(RF_array[:,2]==Low_RF)[0][0]

    #print(str(len(RF_Best_Index[0]))+","+str(len(RF_Low_Index)[0]))
    print("Best Scoring Random Forest: " + str(Best_RF)[0:6] + "  with " + str(RF_array[RF_Best_Index][0]) + " trees")
    print("Lowest MSE Random Forest: " + str(Low_RF)[0:6] + "  with " + str(RF_array[RF_Low_Index][0]) + " trees")
    print("-----")

def XT_analysis():

    XT_array = np.array(XT_results)
    Best_XT = np.max(XT_array[:,1])
    Low_XT = np.min(XT_array[:,2])

    XT_Best_Index = np.where(XT_array[:,1]==Best_XT)[0][0]
    XT_Low_Index = np.where(XT_array[:,2]==Low_XT)[0][0]

    #print(str(len(XT_Best_Index[0]))+","+str(len(XT_Low_Index)[0]))
    print("Best Scoring ExtraTrees: " + str(Best_XT)[0:6] + "  with " + str(XT_array[XT_Best_Index][0]) + " trees")
    print("Lowest MSE ExtraTrees: " + str(Low_XT)[0:6] + "  with " + str(XT_array[XT_Low_Index][0]) + " trees")
    print("-----")

def GB_analysis():

    GB_array = np.array(GB_results)
    Best_GB = np.max(GB_array[:,3])
    Low_GB = np.min(GB_array[:,4])

    GB_Best_Index = np.where(GB_array[:,3]==Best_GB)[0][0]
    GB_Low_Index = np.where(GB_array[:,4]==Low_GB)[0][0]

    #print(str(len(GB_Best_Index[0]))+","+str(len(GB_Low_Index)[0]))
    print("Best Scoring Gradient Boosted Forest: " + str(Best_GB)[0:6] + "  with " + str(GB_array[GB_Best_Index][0]) + " trees & " + str(GB_array[GB_Best_Index][2]) + " rate & " + str(GB_array[GB_Best_Index][1]) + " depth")
    print("Lowest MSE Gradient Boosted Forest: " + str(Low_GB)[0:6] + "  with " + str(GB_array[GB_Low_Index][0]) + " trees & " + str(GB_array[GB_Low_Index][2]) + " rate & " + str(GB_array[GB_Low_Index][1]) + " depth")
    print("-----")

def KN_analysis():

    KN_array = np.array(KN_results)
    Best_KN = np.max(KN_array[:,3].astype(float))
    Low_KN = np.min(KN_array[:,4].astype(float))

    KN_Best_Index = np.where(KN_array[:,3].astype(float)==Best_KN)[0][0]
    KN_Low_Index = np.where(KN_array[:,4].astype(float)==Low_KN)[0][0]

    print("Best Scoring k-Nearest Neighbours: " + str(Best_KN)[0:6] + "  with " + str(KN_array[KN_Best_Index][2]) + " neighbours & " + str(KN_array[KN_Best_Index][0]) + "leafs & " + str(KN_array[KN_Best_Index][1]) + " weights")
    print("Lowest MSE k-Nearest Neighbours: " + str(Low_KN)[0:6] + "  with " + str(KN_array[KN_Low_Index][2]) + " neighbours & " + str(KN_array[KN_Low_Index][0]) + "leafs & " + str(KN_array[KN_Low_Index][1]) + " weights")
    print("-----")

def SV_analysis():

    SV_array = np.array(SV_results)
    Best_SV = np.max(SV_array[:,3].astype(float))
    Low_SV = np.min(SV_array[:,4].astype(float))

    SV_array1 = np.array(SVP_results)
    Best_SV1 = np.max(SV_array1[:,3].astype(float))
    Low_SV1 = np.min(SV_array1[:,4].astype(float))

    SV_Best_Index = np.where(SV_array[:,3].astype(float)==Best_SV)[0][0]
    SV_Low_Index = np.where(SV_array[:,4].astype(float)==Low_SV)[0][0]

    SV_Best_Index1 = np.where(SV_array1[:,3].astype(float)==Best_SV1)[0][0]
    SV_Low_Index1 = np.where(SV_array1[:,4].astype(float)==Low_SV1)[0][0]

    if Best_SV > Best_SV1:
        print("Best Scoring Support Vector Machine: " + str(Best_SV)[0:6] + "  with " + str(SV_array[SV_Best_Index][0]) + " nu & " + str(SV_array[SV_Best_Index][1]) + " C & " + str(SV_array[SV_Best_Index][2] + " kernel"))
    else:
        print("Best Scoring Support Vector Machine: " + str(Best_SV1)[0:6] + "  with " + str(SV_array1[SV_Best_Index1][0]) + " nu & " + str(SV_array1[SV_Best_Index1][1]) + " C & Polynomial kernel of Degree " + str(SV_array1[SV_Best_Index1][2]))
    
    if Low_SV < Low_SV1:
        print("Lowest MSE Support Vector Machine: " + str(Low_SV)[0:6] + "  with " + str(SV_array[SV_Low_Index][0]) + " nu & " + str(SV_array[SV_Best_Index][1]) + " C & " + str(SV_array[SV_Best_Index][2] + " kernel"))
    else:
        print("Lowest MSE Support Vector Machine: " + str(Low_SV1)[0:6] + "  with " + str(SV_array1[SV_Low_Index1][0]) + " nu & " + str(SV_array1[SV_Low_Index1][1]) + " C & Polynomial kernel of Degree " + str(SV_array1[SV_Low_Index1][2]))
    
    print("-----")

def MP_analysis():

    MP_array = np.array(MP_results)
    Best_MP = np.max(MP_array[:,2].astype(float))
    Low_MP = np.min(MP_array[:,3].astype(float))

    MP_Best_Index = np.where(MP_array[:,2].astype(float)==Best_MP)[0]
    MP_Low_Index = np.where(MP_array[:,3].astype(float)==Low_MP)[0]

    print("Best Scoring Multi-Layer Perceptron: " + str(Best_MP)[0:6])
    for index in MP_Best_Index:
        print("with Neurons/Layers: " + str(MP_array[index][0]) + " & Activation Function " + str(MP_array[index][1]))
        print(index)
    print("Lowest MSE Multi-Layer Perceptron: " + str(Low_MP)[0:6])
    for index in MP_Low_Index:
        print("with Neurons/Layers: " + str(MP_array[index][0]) + " & Activation Function " + str(MP_array[index][1]))
        print(index)
    print("-----")

RF_analysis(), XT_analysis(), GB_analysis(), KN_analysis(), SV_analysis(), MP_analysis()