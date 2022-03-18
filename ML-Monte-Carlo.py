#ML Model Monte-Carlo sensitivity analysis Script

#importing libraries
import numpy as np
from sklearn.ensemble import ExtraTreesRegressor #ML model
import xlrd #Excel reading
import matplotlib.pyplot as plt #Plotting
import openpyxl #Excel writing
#----------------------------------------------------------------------------------------
#INPUT VALUES (FROM CFD)
Volume =170    #cloud volume (m3)
Fraction = 0.06042  #Volume fraction of H2 (0>H2>1)

MaxSpreadCloud = 0.2     #Maximum +- % variability of Cloud Size (0<x<1)
MaxSpreadFraction = 0.2     #Maximum +- % variability of Volume Fraction (0<x<1)
n_Samples = 1000    #Number of random samples to use in simulation
#----------------------------------------------------------------------------------------
#Numpy reproducibility seed
np.random.seed(0)

#Distance range
distanceRange = range(5, 100 +1, 1)
distanceRange = np.array(distanceRange)

#importing dataset
DataBook1 = xlrd.open_workbook("ExperimentData.xls")
DataSheet1 = DataBook1.sheet_by_index(0)

featureData = []        #array for H2 conc%, Volume, Distance - feature data list (2D)
pressureData = []       #array for result/pressure data list
for i in range(2, 88 +1, 1):
    featureData.append([DataSheet1.cell_value(i, 1),DataSheet1.cell_value(i, 2),DataSheet1.cell_value(i, 3)])    #adding all feature and pressure value data into arrays from excel document
    pressureData.append(DataSheet1.cell_value(i, 4))

#Generating normal distribution array of cloud volumes
VolumeDistribution = np.random.normal(loc=Volume, scale=MaxSpreadCloud*Volume, size=n_Samples)
#Generating uniform distribution array of H2 volume fractions
H2Distribution = np.random.uniform(low=Fraction-Fraction*MaxSpreadFraction, high=Fraction+Fraction*MaxSpreadFraction, size=n_Samples)

#Creating ML Model for predicting
Estimator = ExtraTreesRegressor(n_estimators=290, criterion="mse", n_jobs=-1, min_samples_leaf=1).fit(featureData, pressureData)

#Empty distance curve list (2D)
DistanceCurveList = []

#looping through all Volume/Fraction Distribution samples
for i in range(n_Samples):
    DistanceCurve = []

    for distance in distanceRange:  #Looping through every distance value for prediction
        DistanceCurve.append(Estimator.predict([[H2Distribution[i]*100, VolumeDistribution[i], distance]]))
    
    print(i)
    DistanceCurveList.append(DistanceCurve) #saving 1D pressure/distance prediction list to 2D list of all predictions

#Plotting pressure/distance curves
plt.figure("Overpressure Vs. Distance")
plt.xlabel("Distance (m)")
plt.ylabel("OverPressure (kPa)")
plt.grid()

for i in range(n_Samples):
    plt.plot(distanceRange, DistanceCurveList[i])

#Saving results to excel sheet
wb1 = openpyxl.Workbook()
s1 = wb1.active
s1.title = "Monte-Carlo Results"
#Writing labels to excel sheet
s1.cell(row=2,column=1,value="Distance")
s1.cell(row=2,column=2,value="Volume")
s1.cell(row=1,column=2,value="Fraction")


for i in range(len(distanceRange)):
    s1.cell(row=3+i,column=1,value=float(distanceRange[i])) #Writing distance list

for i in range(n_Samples):
    s1.cell(row=1,column=3+i).value = H2Distribution[i] #Writing input H2 volume fraction
    s1.cell(row=2,column=3+i).value = VolumeDistribution[i] #Writing input cloud volume

    for j in range(len(distanceRange)):
        s1.cell(row=3+j,column=3+i).value = float(DistanceCurveList[i][j])  #Writing prediction data

#Saving excel sheet - WILL OVERWRITE!!!
wb1.save("MC-results-ML.xlsx")

#Show Plot
plt.show()