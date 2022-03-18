import numpy as np
import matplotlib.pyplot as plt
from numpy.lib.stride_tricks import _maybe_view_as_subclass
import openpyxl
from sklearn.ensemble import ExtraTreesRegressor
import xlrd

w = openpyxl.load_workbook("MC-results-ML.xlsx")
s = w.active

volume=170
fraction = 0.06042 * 100

distanceList = [*range(5,100+1,1)]

def Model():
    #importing dataset
    DataBook1 = xlrd.open_workbook("ExperimentData.xls")
    DataSheet1 = DataBook1.sheet_by_index(0)

    featureData = []        #array for H2 conc%, Volume, Distance - feature data list (2D)
    pressureData = []       #array for result/pressure data list
    for i in range(2, 88 +1, 1):
        featureData.append([DataSheet1.cell_value(i, 1),DataSheet1.cell_value(i, 2),DataSheet1.cell_value(i, 3)])    #adding all feature and pressure value data into arrays from excel document
        pressureData.append(DataSheet1.cell_value(i, 4))

    model = ExtraTreesRegressor(n_estimators=290, n_jobs=-1).fit(featureData,pressureData)

    result = []

    for distance in distanceList:
        result.append(model.predict([[fraction,volume,distance]]))
    
    return result

results = Model()

MC_results = []
MC_input = []

for column in s.iter_cols(min_row=3, max_row=98, min_col=3, max_col=1002, values_only=True):
    MC_results.append(column)

for column in s.iter_cols(min_row=1, max_row=2, min_col=3, max_col=1002, values_only=True):
    MC_input.append(column)

MC_results = np.array(MC_results)

max = np.where(MC_results == np.max(MC_results))[0]
min = np.where(MC_results == np.min(MC_results))[0]

plt.figure("ML-Important MC results")
plt.plot(distanceList, results)
plt.plot(distanceList, np.transpose(MC_results[max]), linestyle = "dashed")
plt.plot(distanceList, np.transpose(MC_results[min]), linestyle = "dashed")
plt.xlabel("Distance (m)")
plt.ylabel ("Overpressure (kPa)")
plt.legend(["Most Probable","Maximum","Minimum"])

plt.grid()
plt.show()