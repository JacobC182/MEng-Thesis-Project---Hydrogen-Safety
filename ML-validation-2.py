#ML tuning and performance validation script
#importing libaries
from numpy.core.numeric import cross
from numpy.random import RandomState
from openpyxl.workbook.workbook import Workbook
from scipy.sparse.construct import rand
import xlrd
import numpy as np
from sklearn.ensemble import RandomForestRegressor, ExtraTreesRegressor, GradientBoostingRegressor
from sklearn.neighbors import KNeighborsRegressor
from sklearn.svm import NuSVR
from sklearn.neural_network import MLPRegressor
import matplotlib.pyplot as plt
from sklearn.model_selection import cross_val_score, GridSearchCV, train_test_split
from sklearn.preprocessing import StandardScaler
import openpyxl

#making distance list
distanceRange = range(2, 100 +1, 1)
distanceRange = np.array(distanceRange)

#importing dataset
DataBook1 = xlrd.open_workbook("ExperimentData.xls")
DataSheet1 = DataBook1.sheet_by_index(0)

featureData = []        #array for H2 conc%, Volume, Distance - feature data list (2D)
pressureData = []       #array for result/pressure data list
for i in range(2, 88 +1, 1):
    featureData.append([DataSheet1.cell_value(i, 1),DataSheet1.cell_value(i, 2),DataSheet1.cell_value(i, 3)])    #adding all feature and pressure value data into arrays from excel document
    pressureData.append(DataSheet1.cell_value(i, 4))

#Doing the validation


RF_score_means = []
XT_score_means = []
GB_score_means = []
KN_score_means = []
SV_score_means = []
MP_score_means = []

RF_mse_means = []
XT_mse_means = []
GB_mse_means = []
KN_mse_means = []
SV_mse_means = []
MP_mse_means = []

def MLPgrid():

    def powerset(fullset):
        listsub = list(fullset)
        subsets = []
        for i in range(2**len(listsub)):
            subset = []
            for k in range(len(listsub)):            
                if i & 1<<k:
                    subset.append(listsub[k])
            subsets.append(subset)        
        return subsets

    subsets = powerset(set([2,3,4,5,6,7,8]))
    subsets2 = subsets

    for subset in subsets:
        if len(subset) > 4:
            subsets2.remove(subset)

    subsets2.pop(0)
    from itertools import permutations
    subsets3 = []

    for subset in subsets2:
        for x in permutations(subset):
            subsets3.append(x)

    return subsets3

layerGrid = MLPgrid()

def MSE(predicted, true):
    if len(predicted) ==  len(true):
        mse = 0
        for i in range(len(predicted)):
            mse = mse + (true[i] - predicted[i])**2

        return mse * (1/len(predicted))
    else:
        raise Exception("Must be equal number of Predictions and Observations!")

activ = ["relu","tanh","logistic"]

Workbook1 = openpyxl.Workbook()
s = Workbook1.active
counter=1

for j in range(len(activ)):
    for k in range(len(layerGrid)):

        MP_score = []

        MP_mse = []
        for i in range(10):
            xtrain, xtest, ytrain, ytest = train_test_split(featureData, pressureData, test_size=0.1, random_state=i*5)   #Split Test/Train datasets

            MLP = MLPRegressor(hidden_layer_sizes=layerGrid[k],solver="lbfgs", activation=activ[j]).fit(xtrain, ytrain)
            MP_score.append(MLP.score(xtest, ytest))

            MP_mse.append(MSE(MLP.predict(xtest), ytest))

        s.cell(row=counter,column=1,value=str(layerGrid[k]))
        s.cell(row=counter,column=2,value=activ[j])

        s.cell(row=counter,column=3,value=float(np.mean(MP_score)))
        s.cell(row=counter,column=4,value=float(np.mean(MP_mse)))
        counter += 1
        print(str(k) + "-------------------------------")

def PrintOut():
    print("--------------------------------------")
    print("Random Forest:")
    print("Mean Score: " + str(np.mean(RF_score))[0:6] + "  Score Std: " + str(np.std(RF_score))[0:6])
    print("Mean MSE: " + str(np.mean(RF_mse))[0:6] + "    MSE Std: " + str(np.std(RF_mse))[0:6])
    print("-")
    print("Extremely Randomised Trees:")
    print("Mean Score: " + str(np.mean(XT_score))[0:6] + "  Score Std: " + str(np.std(XT_score))[0:6])
    print("Mean MSE: " + str(np.mean(XT_mse))[0:6] + "    MSE Std: " + str(np.std(XT_mse))[0:6])
    print("-")
    print("Gradient Boosting Trees:")
    print("Mean Score: " + str(np.mean(GB_score))[0:6] + "  Score Std: " + str(np.std(GB_score))[0:6])
    print("Mean MSE: " + str(np.mean(GB_mse))[0:6] + "    MSE Std: " + str(np.std(GB_mse))[0:6])
    print("-")
    print("k-Nearest Neighbours:")
    print("Mean Score: " + str(np.mean(KN_score))[0:6] + "  Score Std: " + str(np.std(KN_score))[0:6])
    print("Mean MSE: " + str(np.mean(KN_mse))[0:6] + "    MSE Std: " + str(np.std(KN_mse))[0:6])
    print("-")
    print("Support Vector Machine:")
    print("Mean Score: " + str(np.mean(SV_score))[0:6] + "  Score Std: " + str(np.std(SV_score))[0:6])
    print("Mean MSE: " + str(np.mean(SV_mse))[0:6] + "    MSE Std: " + str(np.std(SV_mse))[0:6])
    print("-")
    print("Multi-Layer Perceptron:")
    print("Mean Score: " + str(np.mean(MP_score))[0:6] + "  Score Std: " + str(np.std(MP_score))[0:6])
    print("Mean MSE: " + str(np.mean(MP_mse))[0:6] + "    MSE Std: " + str(np.std(MP_mse))[0:6])

#PrintOut()

def save():
    Workbook1 = openpyxl.Workbook()
    s = Workbook1.active
    s.title = "CV-results"
    s.cell(row=1,column=1,value="GBscore")
    s.cell(row=1,column=2,value="GB-mse")
    #s.cell(row=1,column=3,value="XTscore")
    #s.cell(row=1,column=4,value="XT-mse")

    for i in range(len(RF_score_means)):
        s.cell(row=2+i,column=1,value=float(RF_score_means[i]))
        s.cell(row=2+i,column=2,value=float(RF_mse_means[i]))
        #s.cell(row=2+i,column=3,value=float(XT_score_means[i]))
        #s.cell(row=2+i,column=4,value=float(XT_mse_means[i]))

    Workbook1.save("CV-Results-GF.xlsx")

#save()

Workbook1.save("CV-results-MLP2.xlsx")